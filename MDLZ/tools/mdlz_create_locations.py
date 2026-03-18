import os
import pandas as pd
import openpyxl
import time
import json
from wiliot_api.platform.platform import LocationType, EntityType

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(TOOLS_DIR)
DATA_DIR = os.path.join(PROJECT_DIR, ".tmp", "mdlz")
RESULTS_DIR = os.path.join(DATA_DIR, "results")
BRANCH_CROSSDOCK_EXCEL = os.path.join(DATA_DIR, "branch_crossdock_match_results.xlsx")
STORE_EXCEL = os.path.join(DATA_DIR, "store_match_results.xlsx")

# Column indices (1-based) in the "Store Match Results" sheet
STORE_COL_ECC = 1
STORE_COL_MATCH_TYPE = 12
STORE_COL_PLATFORM_ID = 13
STORE_COL_PLATFORM_NAME = 14
STORE_COL_PLATFORM_LAT = 15
STORE_COL_PLATFORM_LNG = 16
STORE_COL_DISTANCE_KM = 17
STORE_COL_PLATFORM_LOC_TYPE = 18


def _save_results(results, filename):
    os.makedirs(RESULTS_DIR, exist_ok=True)
    path = os.path.join(RESULTS_DIR, filename)
    with open(path, "w") as f:
        json.dump(results, f, indent=2)
    success = sum(1 for r in results if r["status"] == "success")
    failed = len(results) - success
    print(f"\nDone. {success} succeeded, {failed} failed.")
    print(f"Results saved to {path}")
    return results


def create_parent_branch_locations(pc, limit=None):
    """
    Creates platform locations and Staging zones for all PARENT BRANCH
    entries with Match_Type == 'no_match'.

    :param pc: An already-instantiated PlatformClient
    """
    df = pd.read_excel(BRANCH_CROSSDOCK_EXCEL)
    parent_no_match = df[(df["Type"] == "PARENT BRANCH") & (df["Match_Type"] == "no_match")]
    if limit is not None:
        parent_no_match = parent_no_match.head(limit)
    print(f"Found {len(parent_no_match)} PARENT BRANCH locations to process\n")

    results = []

    for _, row in parent_no_match.iterrows():
        depot_num = str(row["Depot_#"])
        name = row["Depot_Name"].strip()
        street = str(row["Street"]).strip() if pd.notna(row["Street"]) else ""
        city = str(row["City"]).strip()
        state = str(row["State"]).strip()
        zipcode = str(int(row["Zip"])) if pd.notna(row["Zip"]) else ""
        lat = float(row["Latitude"])
        lng = float(row["Longitude"])

        address = f"{street}, {city}, {state} {zipcode}".strip(", ")

        print(f"[{depot_num}] Creating location: {name}")
        try:
            location = pc.create_location(
                location_type=LocationType.SITE,
                name=name,
                lat=lat,
                lng=lng,
                address=address,
                city=city,
                country="US",
                is_soft_asset_create=False,
            )
            location_id = location["id"]
            print(f"  Location created: {location_id}")

            pc.set_keys_values_for_entities(
                entity_type=EntityType.LOCATION,
                entity_ids=[location_id],
                keys_values={
                    "locationType": "DC",
                    "locationNumber": depot_num,
                },
                overwrite_existing=True,
            )
            print(f"  Labels set: locationType=DC, locationNumber={depot_num}")

            zone = pc.create_zone(name="Staging", location_id=location_id)
            zone_id = zone["id"]
            print(f"  Zone created: {zone_id}")

            pc.set_keys_values_for_entities(
                entity_type=EntityType.ZONE,
                entity_ids=[zone_id],
                keys_values={"template_name": "DC_Staging"},
                overwrite_existing=True,
            )
            print(f"  Zone label set: template_name=DC_Staging")

            results.append({
                "depot_num": depot_num,
                "depot_name": name,
                "location_id": location_id,
                "zone_id": zone_id,
                "status": "success",
            })

        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({
                "depot_num": depot_num,
                "depot_name": name,
                "location_id": None,
                "zone_id": None,
                "status": f"error: {e}",
            })

        time.sleep(0.5)

    return _save_results(results, "creation_results.json")


def create_crossdock_locations(pc, limit=None):
    """
    Creates platform locations for all CROSSDOCK entries with
    Match_Type == 'no_match' that have a Parent_Branch_Platform_ID.

    Each location gets two labels:
      - locationType = XD
      - ASSOCIATE_DC = <parent branch platform location ID>

    :param pc: An already-instantiated PlatformClient
    """
    df = pd.read_excel(BRANCH_CROSSDOCK_EXCEL)
    crossdocks = df[
        (df["Type"] == "CROSSDOCK")
        & (df["Match_Type"] == "no_match")
        & (df["Parent_Branch_Platform_ID"].notna())
    ]
    if limit is not None:
        crossdocks = crossdocks.head(limit)
    print(f"Found {len(crossdocks)} CROSSDOCK locations to process\n")

    results = []

    for _, row in crossdocks.iterrows():
        depot_num = str(row["Depot_#"])
        name = row["Depot_Name"].strip()
        street = str(row["Street"]).strip() if pd.notna(row["Street"]) else ""
        city = str(row["City"]).strip()
        state = str(row["State"]).strip()
        zipcode = str(int(row["Zip"])) if pd.notna(row["Zip"]) else ""
        lat = float(row["Latitude"])
        lng = float(row["Longitude"])
        parent_platform_id = str(row["Parent_Branch_Platform_ID"])

        address = f"{street}, {city}, {state} {zipcode}".strip(", ")
        location_name = f"{name} - {depot_num}"

        print(f"[{depot_num}] Creating crossdock location: {location_name}")
        try:
            location = pc.create_location(
                location_type=LocationType.SITE,
                name=location_name,
                lat=lat,
                lng=lng,
                address=address,
                city=city,
                country="US",
                is_soft_asset_create=False,
            )
            location_id = location["id"]
            print(f"  Location created: {location_id}")

            pc.set_keys_values_for_entities(
                entity_type=EntityType.LOCATION,
                entity_ids=[location_id],
                keys_values={
                    "locationType": "XD",
                    "ASSOCIATE_DC": parent_platform_id,
                },
                overwrite_existing=True,
            )
            print(f"  Labels set: locationType=XD, ASSOCIATE_DC={parent_platform_id}")

            results.append({
                "depot_num": depot_num,
                "depot_name": name,
                "location_id": location_id,
                "parent_platform_id": parent_platform_id,
                "status": "success",
            })

        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({
                "depot_num": depot_num,
                "depot_name": name,
                "location_id": None,
                "parent_platform_id": parent_platform_id,
                "status": f"error: {e}",
            })

        time.sleep(0.5)

    return _save_results(results, "crossdock_creation_results.json")


def create_store_locations(pc, branch_name, associate_dc_id, limit=None):
    """
    Creates platform locations for all stores with Match_Type == 'no_match'
    under the given Branch_Name. Also checks already-matched stores and
    renames any whose platform name contains the ECC suffix.

    Each location gets three labels:
      - ECC = <store ECC>
      - locationType = store
      - ASSOCIATE_DC = <branch platform location ID>

    Updates the Excel file after each successful creation so the script
    can be safely stopped and re-run.

    :param pc: An already-instantiated PlatformClient
    :param branch_name: Branch_Name to filter stores by
    :param associate_dc_id: Platform location ID of the associated DC
    """
    # Fetch the DC location name from the platform
    dc_location = pc.get_location(associate_dc_id)
    dc_name = dc_location["name"]
    print(f"DC: {dc_name} ({associate_dc_id})")
    print(f"Filtering stores by Branch_Name: '{branch_name}'\n")

    df = pd.read_excel(STORE_EXCEL, sheet_name="Store Match Results")
    branch_stores = df[df["Branch_Name"] == branch_name]

    no_match = branch_stores[branch_stores["Match_Type"] == "no_match"]
    matched = branch_stores[branch_stores["Match_Type"] != "no_match"]

    if limit is not None:
        no_match = no_match.head(limit)
    print(f"Found {len(no_match)} stores to create, {len(matched)} already matched\n")

    # Load the workbook for in-place updates and build ECC -> row lookup
    wb = openpyxl.load_workbook(STORE_EXCEL)
    ws = wb["Store Match Results"]
    ecc_to_row = {}
    for r in range(2, ws.max_row + 1):
        ecc_val = ws.cell(r, STORE_COL_ECC).value
        if ecc_val is not None:
            ecc_to_row[str(int(ecc_val))] = r

    results = []

    # --- Fix names of already-matched stores that have ECC in the name ---
    renamed = 0
    for _, row in matched.iterrows():
        platform_id = row.get("Platform_ID")
        if pd.isna(platform_id):
            continue
        ecc = str(int(row["ECC"]))

        try:
            location = pc.get_location(platform_id)
            current_name = location.get("name", "")
            ecc_suffix = f" - {ecc}"
            if current_name.endswith(ecc_suffix):
                new_name = current_name[: -len(ecc_suffix)]
                location["name"] = new_name
                pc.update_location(location)
                renamed += 1
                print(f"  RENAMED: {current_name} -> {new_name}")
        except Exception as e:
            print(f"  RENAME ERROR for {platform_id}: {e}")

        time.sleep(0.3)

    if renamed:
        print(f"\nRenamed {renamed} existing locations\n")

    # --- Create new store locations ---
    for _, row in no_match.iterrows():
        ecc = str(int(row["ECC"]))
        store_name = str(row["Store_Name"]).strip()
        street = str(row["Street"]).strip() if pd.notna(row["Street"]) else ""
        city = str(row["City"]).strip()
        state = str(row["State"]).strip()
        zipcode = str(row["ZIP"]).strip() if pd.notna(row["ZIP"]) else ""
        lat = float(row["Store_Lat"])
        lng = float(row["Store_Lng"])

        address = f"{street}, {city}, {state} {zipcode}".strip(", ")

        print(f"[{ecc}] Creating store location: {store_name}")
        try:
            location = pc.create_location(
                location_type=LocationType.SITE,
                name=store_name,
                lat=lat,
                lng=lng,
                address=address,
                city=city,
                country="US",
                is_soft_asset_create=False,
            )
            location_id = location["id"]
            print(f"  Location created: {location_id}")

            pc.set_keys_values_for_entities(
                entity_type=EntityType.LOCATION,
                entity_ids=[location_id],
                keys_values={
                    "ECC": ecc,
                    "locationType": "store",
                    "parentBranch": dc_name,
                    "ASSOCIATE_DC": associate_dc_id,
                },
                overwrite_existing=True,
            )
            print(f"  Labels set: ECC={ecc}, locationType=store, parentBranch={dc_name}, ASSOCIATE_DC={associate_dc_id}")

            # Update the Excel row
            excel_row = ecc_to_row.get(ecc)
            if excel_row:
                ws.cell(excel_row, STORE_COL_MATCH_TYPE).value = "name+coords"
                ws.cell(excel_row, STORE_COL_PLATFORM_ID).value = location_id
                ws.cell(excel_row, STORE_COL_PLATFORM_NAME).value = store_name
                ws.cell(excel_row, STORE_COL_PLATFORM_LAT).value = lat
                ws.cell(excel_row, STORE_COL_PLATFORM_LNG).value = lng
                ws.cell(excel_row, STORE_COL_DISTANCE_KM).value = 0.0
                ws.cell(excel_row, STORE_COL_PLATFORM_LOC_TYPE).value = "site"
                wb.save(STORE_EXCEL)
                print(f"  Excel updated: row {excel_row}")

            results.append({
                "ecc": ecc,
                "store_name": store_name,
                "location_id": location_id,
                "associate_dc_id": associate_dc_id,
                "status": "success",
            })

        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({
                "ecc": ecc,
                "store_name": store_name,
                "location_id": None,
                "associate_dc_id": associate_dc_id,
                "status": f"error: {e}",
            })

        time.sleep(0.5)

    os.makedirs(RESULTS_DIR, exist_ok=True)
    results_path = os.path.join(RESULTS_DIR, "store_creation_results.json")
    with open(results_path, "w") as f:
        json.dump(results, f, indent=2)

    success = sum(1 for r in results if r["status"] == "success")
    failed = len(results) - success
    print(f"\nDone. {success} created, {failed} failed. {renamed} renamed.")
    print(f"Results saved to {results_path}")
    return results


def create_trailer_locations(pc, excel_path, dc_location_id, template_name, sheet_name="LIST", limit=None):
    """
    Creates TRANSPORTER locations for each trailer in the given Excel file.
    Each location gets a zone and labels for parentBranch and ASSOCIATE_DC.

    :param pc: An already-instantiated PlatformClient
    :param excel_path: Path to the Excel file with trailer data
    :param dc_location_id: Platform location ID of the associated DC
    :param template_name: Zone template name (e.g. 'Standard_truck')
    :param sheet_name: Sheet name to read from (default 'LIST')
    """
    dc_location = pc.get_location(dc_location_id)
    branch_name = dc_location["name"]
    print(f"DC: {branch_name} ({dc_location_id})")
    print(f"Zone template: {template_name}\n")

    resolved_path = os.path.join(DATA_DIR, excel_path) if not os.path.isabs(excel_path) else excel_path
    wb = openpyxl.load_workbook(resolved_path)
    ws = wb[sheet_name]

    max_row = ws.max_row
    if limit is not None:
        max_row = min(1 + limit, ws.max_row)

    results = []

    for r in range(2, max_row + 1):
        unit_num = str(ws.cell(r, 1).value).strip()
        location_name = f"Truck{unit_num}"

        print(f"[{unit_num}] Creating transporter: {location_name}")
        try:
            location = pc.create_location(
                location_type=LocationType.TRANSPORTER,
                name=location_name,
                is_soft_asset_create=False,
            )
            location_id = location["id"]
            print(f"  Location created: {location_id}")

            pc.set_keys_values_for_entities(
                entity_type=EntityType.LOCATION,
                entity_ids=[location_id],
                keys_values={
                    "parentBranch": branch_name,
                    "ASSOCIATE_DC": dc_location_id,
                },
                overwrite_existing=True,
            )
            print(f"  Labels set: parentBranch={branch_name}, ASSOCIATE_DC={dc_location_id}")

            zone = pc.create_zone(name=unit_num, location_id=location_id)
            zone_id = zone["id"]
            print(f"  Zone created: {zone_id}")

            pc.set_keys_values_for_entities(
                entity_type=EntityType.ZONE,
                entity_ids=[zone_id],
                keys_values={"template_name": template_name},
                overwrite_existing=True,
            )
            print(f"  Zone label set: template_name={template_name}")

            results.append({
                "unit_num": unit_num,
                "location_name": location_name,
                "location_id": location_id,
                "zone_id": zone_id,
                "status": "success",
            })

        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({
                "unit_num": unit_num,
                "location_name": location_name,
                "location_id": None,
                "zone_id": None,
                "status": f"error: {e}",
            })

        time.sleep(0.5)

    return _save_results(results, "trailer_creation_results.json")


def add_ecc_labels(pc, sheet="Store Match Results", match_type="name+coords", limit=None):
    """
    Reads the given sheet from the store Excel file and adds ECC labels to
    matched platform locations.

    For 'Store Match Results': filters by Match_Type, uses ECC and Platform_ID.
    For 'No ECC Analysis': filters by Match_Status == 'MATCHED', uses
    Matched_ECC and Platform_ID.

    :param pc: An already-instantiated PlatformClient
    :param sheet: Sheet name to read from
    :param match_type: Match_Type value to filter (only for Store Match Results)
    """
    BATCH_SIZE = 50
    df = pd.read_excel(STORE_EXCEL, sheet_name=sheet)

    if sheet == "No ECC Analysis":
        matched = df[df["Match_Status"] == "MATCHED"].copy()
        matched["ECC"] = matched["Matched_ECC"].astype(int).astype(str)
        print(f"Found {len(matched)} MATCHED locations in '{sheet}'\n")
    else:
        matched = df[df["Match_Type"] == match_type].copy()
        matched["ECC"] = matched["ECC"].astype(str)
        print(f"Found {len(matched)} locations with Match_Type='{match_type}'\n")

    if limit is not None:
        matched = matched.head(limit)

    success = 0
    skipped = 0
    failed = 0
    total = len(matched)

    for start in range(0, total, BATCH_SIZE):
        batch = matched.iloc[start : start + BATCH_SIZE]
        for _, row in batch.iterrows():
            location_id = row["Platform_ID"]
            ecc_value = row["ECC"]
            count = success + skipped + failed
            try:
                existing = pc.get_entity_keys_values(
                    entity_type=EntityType.LOCATION,
                    entity_id=location_id,
                    key="ECC",
                )
                if existing:
                    existing_ecc = existing[0]["label"]["value"]
                    skipped += 1
                    print(f"  [{count + 1}/{total}] SKIP {location_id} — already has ECC={existing_ecc}")
                    continue

                pc.set_keys_values_for_entities(
                    entity_type=EntityType.LOCATION,
                    entity_ids=[location_id],
                    keys_values={"ECC": ecc_value},
                    overwrite_existing=True,
                )
                success += 1
                print(f"  [{count + 1}/{total}] {location_id} -> ECC={ecc_value}")
            except Exception as e:
                failed += 1
                print(f"  [{count + 1}/{total}] ERROR {location_id}: {e}")

        time.sleep(0.3)

    print(f"\nDone. {success} created, {skipped} skipped (existing), {failed} failed out of {total}.")
